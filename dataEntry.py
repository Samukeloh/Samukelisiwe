from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")




file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="PhoneNumber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    
    file.save('Backened_data.xlsx')




def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save('Backened_data.xlsx')

    messagebox.showinfo('info','details added')

    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)
    
    

   

def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)





#icon


#heading
Label(root,text="Please fill ou this Entry form",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text="Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Contact No",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="Address",font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#Entry
nameValue=StringVar()
contactValue=StringVar() 
AgeValue=StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactEntry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

#gender
gender_combobox = Combobox(root,values=['Male','Female'],font='arial 11',state='r',width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Male')



addressEntry = Text(root,width=50,height=4,bd=4)



nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)


Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)









root.mainloop()
